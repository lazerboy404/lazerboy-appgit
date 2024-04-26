import telegram
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes, MessageHandler, filters
import openpyxl
import requests
from io import BytesIO

# Diccionario para asociar emojis a encabezados
emojis_encabezados = {
    'SITIO': '🌐',
    'HOSTNAME': '🖥️',
    'SECTOR': '🏢',
    'ASIGNACION': '📋',
    'PROYECCION': '🔍',
    'MODELO': '💻',
    'SERIE': '🔢',
    'CAMS': '🖥️',
    'RAM': '🖥️',
    'CAPACIDAD HD': '💽',
    'FRECUENCIA': '📡',
    'IP VOZ': '📞',
    'MAC VOZ': '🔒',
    'IP DATOS': '💻',
    'MAC DATOS': '🔒'
}

# Función para leer el archivo Excel desde una URL
def leer_excel_desde_url(url, palabra_busqueda):
    response = requests.get(url)
    workbook = openpyxl.load_workbook(BytesIO(response.content))
    encabezados_datos = []

    # Iterar sobre todas las hojas
    for hoja in workbook.sheetnames:
        current_sheet = workbook[hoja]
        current_encabezados = [celda.value for celda in current_sheet[1]]  # Obtener encabezados de la hoja

        # Iterar sobre todas las filas de la hoja
        for fila in range(2, current_sheet.max_row + 1):  # Comenzar desde la segunda fila para evitar los encabezados
            fila_datos = [current_sheet.cell(row=fila, column=columna).value for columna in range(1, current_sheet.max_column + 1)]
            if any(palabra_busqueda.lower() in str(valor).lower() for valor in fila_datos):
                encabezados_datos.append((current_encabezados, fila_datos))  # Agregar encabezados y datos

    return encabezados_datos

# Función para manejar el comando /buscar
async def buscar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Obtener la palabra de búsqueda del mensaje del usuario
    palabra_busqueda = ' '.join(context.args).lower()
    if not palabra_busqueda:
        await context.bot.send_message(chat_id=update.effective_chat.id, text="Por favor, proporciona una palabra para buscar después del comando.")
        return
    
    # URL del archivo Excel
    url_archivo = "https://myawsbucketlazerboy.s3.us-east-2.amazonaws.com/bd+bot.xlsx"
    encabezados_datos = leer_excel_desde_url(url_archivo, palabra_busqueda)
    
    mensaje = 'Resultados para "{}":\n\n'.format(palabra_busqueda)
    if encabezados_datos:
        for encabezados, fila_datos in encabezados_datos:
            for encabezado, valor in zip(encabezados, fila_datos):
                emoji_encabezado = emojis_encabezados.get(encabezado, '')  # Obtener emoji del encabezado
                mensaje += '{} {}: {}\n'.format(emoji_encabezado, encabezado, valor)
            mensaje += '\n'  # Agregar espacio entre cada conjunto de resultados
    else:
        mensaje += "No se encontraron resultados."

    await context.bot.send_message(chat_id=update.effective_chat.id, text=mensaje)

# Función para manejar el comando /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Mensaje de bienvenida
    mensaje = "¡Hola! Soy un bot de Telegram. Aquí tienes una lista de los comandos disponibles:\n\n"
    mensaje += "/start - Muestra este mensaje de bienvenida y los comandos disponibles.\n"
    mensaje += "/buscar [palabra] - Busca la palabra en todas las hojas del archivo Excel y muestra los resultados.\n"
    mensaje += "/estadisticas - Muestra estadísticas relevantes sobre el uso del bot."

    await context.bot.send_message(chat_id=update.effective_chat.id, text=mensaje)

# Función para manejar el comando /estadisticas
async def estadisticas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Aquí podrías incluir la lógica para recopilar y mostrar estadísticas sobre el uso del bot
    mensaje = "Estadísticas:\n\n"
    mensaje += "Total de usuarios atendidos: X\n"
    mensaje += "Total de consultas realizadas: Y\n"
    mensaje += "Otra estadística relevante: Z"

    await context.bot.send_message(chat_id=update.effective_chat.id, text=mensaje)

# Configurar el bot con tu token de Telegram
application = ApplicationBuilder().token('7075443977:AAFoKrqTawFuzaeSsnueFE0-tMPAq6uoIJU').build()

# Agregar los manejadores de comandos
application.add_handler(CommandHandler('start', start))
application.add_handler(CommandHandler('buscar', buscar))
application.add_handler(CommandHandler('estadisticas', estadisticas))

# Iniciar el bot
application.run_polling()
